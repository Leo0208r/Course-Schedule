from turtle import color
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, PatternFill, Side
workbook=Workbook()
sheet=workbook.active
sheet.column_dimensions['A'].width=15
sheet.column_dimensions['D'].width=12
Board=[
    ["HOURS","Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"],
    ["7:00-8:00","","","","","","",""],
    ["8:00-9:00","","","","","","",""],
    ["9:00-10:00","","","","","","",""],
    ["10:00-11:00","","","","","","",""],
    ["11:00-12:00","","","","","","",""],
    ["12:00-13:00","","","","","","",""],
    ["13:00-14:00","","","","","","",""],
    ["14:00-15:00","","","","","","",""],
    ["15:00-16:00","","","","","","",""],
    ["16:00-17:00","","","","","","",""],
    ["17:00-18:00","","","","","","",""],
    ["18:00-19:00","","","","","","",""],
    ["19:00-20:00","","","","","","",""],
    ["20:00-21:00","","","","","","",""],
    ["21:00-22:00","","","","","","",""],
]
for i in range(len(Board)):
        for j in range(len(Board[i])):
            sheet.cell(row=i+1, column=j+1, value=Board[i][j])
            sheet.cell(row=i+1, column=j+1).font=Font(bold=True)
            sheet.cell(row=i+1, column=j+1).alignment=Alignment(horizontal="center", vertical="center", wrap_text=True)
            sheet.cell(row=i+1, column=j+1).border=Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            
def validateDay(day):
    days=["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"]
    if day.lower() not in days:
        print("Please enter a valid day of the week.")
        return False
    return True
def validateTime(time):
    try:
        hour, minute = map(int, time.split(":"))
        if 0 <= hour < 24 and minute == 0:
            return True
        else:
            print("Please enter a valid time in 24-hour format (e.g., 14:00).")
            return False
    except ValueError:
        print("Please enter a valid time in 24-hour format (e.g., 14:00).")
        return False
def validateTimeRange(startTime, endTime):
    startHour = int(startTime.split(":")[0])
    endHour = int(endTime.split(":")[0])
    if startHour >= endHour:
        print("End time must be after start time.")
        return False
    return True
def validateChoice(choice):
    if choice.lower() not in ["yes", "no"]:
        print("Please enter 'yes' or 'no'.")
        return False
    return True
    
    
    
def applyCourseToBoard(course, startTime, endTime, day, colour):
    colors={
        "yellow":"FFFF00",
        "greenlight":"90EE90",
        "redlight":"FF7F7F",
        "bluelight":"ADD8E6",
        "salmon":"FFA07A",
        "purplelight":"DDA0DD"
    }
    timeSlots={
        "7:00":1,
        "8:00":2,
        "9:00":3,
        "10:00":4,
        "11:00":5,
        "12:00":6,
        "13:00":7,
        "14:00":8,
        "15:00":9,
        "16:00":10,
        "17:00":11,
        "18:00":12,
        "19:00":13,
        "20:00":14,
        "21:00":15,
        "22:00":16
    }
    dayColumn={
        "monday":2,
        "tuesday":3,
        "wednesday":4,
        "thursday":5,
        "friday":6,
        "saturday":7,
        "sunday":8
    }
    startHour=startTime.split(":")[0]+":00"
    endHour=endTime.split(":")[0]+":00"
    startRow=timeSlots[startHour]
    endRow=timeSlots[endHour]
    column=dayColumn[day.lower()]
    for row in range(startRow, endRow):
        sheet.column_dimensions[sheet.cell(row=1, column=column).column_letter].width=len(course)+2
        sheet.cell(row=row+1, column=column, value=course)
        sheet.cell(row=row+1, column=column).fill=PatternFill(start_color=colors[colour], end_color=colors[colour], fill_type="solid")
    sheet.merge_cells(start_row=startRow+1,start_column=column,end_row=endRow,end_column=column)
    sheet.cell(row=startRow+1, column=column).alignment=Alignment(horizontal="center", vertical="center", wrap_text=True)
    

while (True):            
    course=input("Write the course name: ")
    startTime=input("Write the start time (hour in range (7:00-22:00) for example: 14:00): ")
    endTime=input("Write the end time (hour in range (7:00-22:00) for example: 16:00): ")
    if validateTimeRange(startTime, endTime)==False:
        continue
    if validateTime(startTime)==False or validateTime(endTime)==False:
        continue
    day=input("Write the day of the week: ")
    if validateDay(day)==False:
        continue
    colour=input("Choose a color for the course (yellow, greenlight, redlight, bluelight, salmon, purplelight): ")
    applyCourseToBoard(course, startTime, endTime, day, colour)
    choose=input("Do you want to add another course? (yes/no): ")
    if validateChoice(choose)==False:
        continue
    if choose.lower() == "no":
        break
nameFile=input("Write the file name: ")
workbook.save(f"{nameFile}.xlsx")